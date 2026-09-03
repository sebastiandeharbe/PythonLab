import re
import sys
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Border, Side

# Ruta del README (busca en la raíz donde esté parado el proyecto)
README_PATH = Path("README.md")

# Definir el borde fino estándar (negro)
thin_side = Side(border_style="thin", color="000000")
cell_border = Border(
    left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
)

def find_excel_file(base_dir: Path = Path(".")) -> Path:
    """Busca el único archivo .xlsx en la carpeta raíz del proyecto."""
    excel_files = [
        f
        for f in base_dir.glob("*.xlsx")
        if not f.name.startswith("~$")  # Ignora archivos temporales de Excel
    ]

    if not excel_files:
        raise FileNotFoundError(
            f"No se encontró ningún archivo .xlsx en: {base_dir.resolve()}"
        )

    if len(excel_files) > 1:
        nombres = [f.name for f in excel_files]
        raise ValueError(
            f"Se encontraron múltiples archivos .xlsx ({nombres}). "
            f"Debe haber solo uno."
        )

    return excel_files[0]


def parse_markdown_table(file_path: Path):
    if not file_path.exists():
        raise FileNotFoundError(
            f"No se encontró el archivo README en: {file_path.resolve()}"
        )

    with open(file_path, "r", encoding="utf-8") as f:
        content = f.read()

    entries = []

    # Expresión regular que busca exactamente el patrón de versión:
    # | Versión (dígitos.dígitos...) | Fecha (YYYY MM DD o con puntos) | Dev | Company | Detalle |
    pattern = re.compile(
        r"\|\s*(\d+(?:\.\d+)+)\s*\|\s*(\d{4}[ .]\d{2}[ .]\d{2})\s*\|\s*([^|]*)\|\s*([^|]*)\|\s*([^|\n\r]+)",
        re.MULTILINE,
    )

    for match in pattern.finditer(content):
        version = match.group(1).strip()
        # Normaliza la fecha reemplazando cualquier espacio por punto -> YYYY.MM.DD
        raw_date = match.group(2).strip().replace(" ", ".")
        detalle = match.group(5).strip()

        entries.append((raw_date, version, detalle))

    return entries


def update_excel(entries, excel_path: Path):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Prefijo del año y mes actual en formato "YYYY.MM" (ej: "2026.09")
    current_month_prefix = datetime.now().strftime("%Y.%m")

    # Leer versiones ya existentes en columna B (desde la fila 4)
    existing_versions = set()
    for row in range(4, ws.max_row + 1):
        val = ws.cell(row=row, column=2).value
        if val is not None:
            existing_versions.add(str(val).strip())

    # Filtrar solo versiones del mes actual que no estén cargadas
    new_entries = [
        e
        for e in entries
        if e[1] not in existing_versions and e[0].startswith(current_month_prefix)
    ]

    if not new_entries:
        print(
            f"No hay versiones nuevas del mes en curso ({current_month_prefix}) para ingresar."
        )
        return

    # Insertar filas nuevas en la posición 4
    ws.insert_rows(4, amount=len(new_entries))

    # Escribir valores y aplicar bordes
    for idx, (fecha, version, detalle) in enumerate(new_entries):
        row_num = 4 + idx

        cell_fecha = ws.cell(row=row_num, column=1, value=fecha)
        cell_version = ws.cell(row=row_num, column=2, value=version)
        cell_detalle = ws.cell(row=row_num, column=3, value=detalle)

        # Aplicar el borde a cada celda de la fila insertada
        cell_fecha.border = cell_border
        cell_version.border = cell_border
        cell_detalle.border = cell_border

    wb.save(excel_path)
    print(
        f"¡Listo! Se agregaron {len(new_entries)} versiones con formato de bordes en '{excel_path.name}'."
    )


if __name__ == "__main__":
    try:
        excel_file = find_excel_file()
        print(f"Archivo Excel detectado: {excel_file.name}")
        data = parse_markdown_table(README_PATH)
        update_excel(data, excel_file)
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)